import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io

# === App Configuration ===
st.set_page_config(page_title="Abwesenheits-Eintrager", layout="centered")

# === Title and Instructions ===
st.title("📊 Abwesenheits-Eintrager für Personalstatistik")

st.markdown("""
Willkommen! Folge diesen Schritten:

1. Exportiere die CSV-Datei aus dem Abwesenheitstool.
2. Verwende eine **leere Excel-Vorlage ohne Makros** (z. B. als `.xlsx` gespeichert).
3. Lade beide Dateien unten hoch.
4. Lade anschließend die fertige Excel-Datei herunter.
5. Öffne dein Makro-Original und **kopiere das Tabellenblatt zurück**.

⚠️ **Wichtig:** Diese App übernimmt **keine Makros**. Bitte nicht `.xlsm` hochladen!
""")

# === Uploads ===
uploaded_csv = st.file_uploader("1️⃣ CSV-Datei hochladen", type=["csv"])
uploaded_xlsx = st.file_uploader("2️⃣ Excel-Vorlage hochladen (.xlsx)", type=["xlsx"])
sheet_name = st.text_input("3️⃣ Tabellenblattname", value="März-Mai 24")

# === Auto-run when both files exist ===
if uploaded_csv and uploaded_xlsx:

    with st.spinner("⏳ Verarbeite Dateien..."):

        try:
            # CSV einlesen
            df = pd.read_csv(uploaded_csv, header=1, sep=';')
            result_dict = {}

            for _, row in df.iterrows():
                if pd.isna(row['Lehrperson']):
                    continue

                key = row['Lehrperson']
                vom = pd.to_datetime(row['vom'], dayfirst=True)
                bis = pd.to_datetime(row['bis'], dayfirst=True)

                # TEST-ANPASSUNG: Jahr zurücksetzen
                if vom.year == 2024:
                    vom = vom.replace(year=2023)
                if bis.year == 2024:
                    bis = bis.replace(year=2023)

                result_dict.setdefault(key, []).append([vom, bis])

            # Excel-Datei laden (keine Makros)
            try:
                wb = load_workbook(uploaded_xlsx, data_only=True)
            except Exception as e:
                st.error(f"❌ Fehler beim Öffnen der Excel-Datei: {e}")
                st.stop()

            # Tabellenblatt wählen oder erstellen
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
            else:
                sheet = wb.create_sheet(sheet_name)

            # Datums-Spalten analysieren
            dates_in_sheet = []
            initial_date = None
            for col in range(7, sheet.max_column + 1):
                date_value = sheet.cell(row=2, column=col).value
                day_name = sheet.cell(row=5, column=col).value

                try:
                    if col == 7:
                        initial_date = pd.to_datetime(date_value, format="%d.%m.%Y")
                        if day_name not in ['Sa', 'So']:
                            dates_in_sheet.append((col, initial_date))
                    elif isinstance(date_value, str) and date_value.startswith('=') and initial_date:
                        offset = int(date_value.split('+')[-1])
                        d = initial_date + pd.Timedelta(days=offset)
                        if day_name not in ['Sa', 'So']:
                            dates_in_sheet.append((col, d))
                    else:
                        d = pd.to_datetime(date_value, format="%d.%m.%Y")
                        if day_name not in ['Sa', 'So']:
                            dates_in_sheet.append((col, d))
                except:
                    continue

            # Daten eintragen
            for row_index, (key, periods) in enumerate(result_dict.items(), start=6):
                sheet.cell(row=row_index, column=3, value=key)
                for vom, bis in periods:
                    if vom and bis:
                        for col, date in dates_in_sheet:
                            if date and vom <= date <= bis:
                                sheet.cell(row=row_index, column=col, value='x')

            # Als .xlsx zum Download bereitstellen
            output = io.BytesIO()
            wb.save(output)
            wb.close()
            output.seek(0)

            st.success("✅ Verarbeitung abgeschlossen.")
            st.download_button(
                label="📥 Excel-Datei herunterladen",
                data=output,
                file_name="Personalkennzahlen.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Ein unerwarteter Fehler ist aufgetreten: {e}")
